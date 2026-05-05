"""Build 10 Excel templates với openpyxl.

Run: py giao-trinh/scripts/build-all-xlsx.py
Output: giao-trinh/bieu-mau-xlsx/
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

OUT_DIR = Path(__file__).parent.parent / 'bieu-mau-xlsx'
OUT_DIR.mkdir(exist_ok=True)

# iPMAC brand
BRAND_RED = 'D41B3F'
BRAND_DARK = '1A1A1A'
LIGHT_GREY = 'F5F5F5'
PLACEHOLDER = '999999'
WHITE = 'FFFFFF'

# Reusable styles
header_fill = PatternFill('solid', fgColor=BRAND_RED)
header_font = Font(name='Calibri', size=11, bold=True, color=WHITE)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
subheader_fill = PatternFill('solid', fgColor=LIGHT_GREY)
subheader_font = Font(name='Calibri', size=11, bold=True, color=BRAND_DARK)

cell_font = Font(name='Calibri', size=10)
cell_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
placeholder_font = Font(name='Calibri', size=10, italic=True, color=PLACEHOLDER)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def style_header_row(ws, row_idx, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[row_idx].height = 30


def style_data_cell(cell, italic=False):
    cell.font = placeholder_font if italic else cell_font
    cell.alignment = cell_align
    cell.border = thin_border


def add_title(ws, title, subtitle=None, merge_to='F1'):
    # Extract column letter — accept 'F1' or just 'F'
    col = ''.join(c for c in merge_to if c.isalpha())
    ws['A1'] = title
    ws['A1'].font = Font(name='Calibri', size=18, bold=True, color=BRAND_RED)
    ws.merge_cells(f'A1:{col}1')
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws['A2'] = subtitle
        ws['A2'].font = Font(name='Calibri', size=11, italic=True, color='6B6B6B')
        ws.merge_cells(f'A2:{col}2')
        ws.row_dimensions[2].height = 20


def add_brand_header(ws, tpl_id, buoi):
    ws.oddHeader.left.text = f'iPMAC · Khoá Ready for BA · {tpl_id}'
    ws.oddHeader.left.size = 9
    ws.oddHeader.right.text = buoi
    ws.oddHeader.right.size = 9


def save(wb, filename):
    out = OUT_DIR / filename
    wb.save(str(out))
    print(f'OK {filename}')
    return out


# ============================================================
def build_tpl_001():
    """Skill Matrix — 6 năng lực × 3 mục tiêu + công thức."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Skill Matrix'
    add_brand_header(ws, 'TPL-001', 'Buổi 1')

    add_title(ws, 'TPL-001 — Self-Assessment Skill Matrix',
              'Tự chấm 1-5 cho 6 nhóm năng lực BA. Lưu mỗi quý — so sánh thấy tiến bộ.', 'E1')

    # Hướng dẫn
    ws['A4'] = 'HƯỚNG DẪN CHẤM ĐIỂM'
    ws['A4'].font = subheader_font
    ws['A4'].fill = subheader_fill
    ws.merge_cells('A4:E4')
    rules = [
        ('1', 'Beginner', 'Chỉ nghe qua, chưa từng làm'),
        ('2', 'Novice', 'Đã thử 1-2 lần, cần mentor hướng dẫn'),
        ('3', 'Practitioner', 'Tự làm được nhưng còn chậm'),
        ('4', 'Proficient', 'Tự làm tốt, có thể hướng dẫn người mới'),
        ('5', 'Expert', 'Master, có thể đào tạo team'),
    ]
    for i, (n, name, desc) in enumerate(rules, 5):
        ws.cell(row=i, column=1, value=n).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=2, value=name).font = Font(bold=True)
        ws.cell(row=i, column=3, value=desc)
        ws.merge_cells(f'C{i}:E{i}')
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = thin_border

    # Bảng tự chấm
    ws.cell(row=11, column=1, value='BẢNG TỰ CHẤM').font = subheader_font
    ws.cell(row=11, column=1).fill = subheader_fill
    ws.merge_cells('A11:E11')

    headers = ['#', 'Nhóm năng lực', 'Hôm nay', 'Mục tiêu 6 tháng', 'Mục tiêu 1 năm']
    for c, h in enumerate(headers, 1):
        ws.cell(row=12, column=c, value=h)
    style_header_row(ws, 12, 5)

    skills = [
        'Khơi gợi yêu cầu (phỏng vấn, workshop, observation)',
        'Phân tích & mô hình hoá (BPMN, ERD, Use Case)',
        'Viết tài liệu (BRD, SRS, User Story)',
        'Thiết kế giải pháp (wireframe, mockup, UX)',
        'Trao đổi & validate (meeting, demo, UAT)',
        'Quản lý yêu cầu (tracking, change, traceability)',
    ]
    for i, skill in enumerate(skills, 1):
        row = 12 + i
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=skill)
        # Empty cells for self-assessment (1-5)
        for col in [3, 4, 5]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border

    # Data validation 1-5
    dv = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
    dv.add(f'C13:E18')
    ws.add_data_validation(dv)

    # Conditional formatting — green for high scores
    ws.conditional_formatting.add(
        'C13:E18',
        ColorScaleRule(start_type='num', start_value=1, start_color='FFCCCC',
                       mid_type='num', mid_value=3, mid_color='FFFFCC',
                       end_type='num', end_value=5, end_color='C6EFCE'))

    # Average row
    ws.cell(row=20, column=2, value='Trung bình').font = Font(bold=True)
    for col in [3, 4, 5]:
        cell = ws.cell(row=20, column=col, value=f'=IFERROR(AVERAGE({get_column_letter(col)}13:{get_column_letter(col)}18),"")')
        cell.font = Font(bold=True, color=BRAND_RED)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    ws.cell(row=20, column=2).border = thin_border

    # 4 trụ cột
    ws.cell(row=23, column=1, value='4 TRỤ CỘT NĂNG LỰC (Buổi 2)').font = subheader_font
    ws.cell(row=23, column=1).fill = subheader_fill
    ws.merge_cells('A23:E23')

    headers2 = ['Trụ cột', 'Hôm nay', 'Mục tiêu 1 năm', 'Plan học', 'Note']
    for c, h in enumerate(headers2, 1):
        ws.cell(row=24, column=c, value=h)
    style_header_row(ws, 24, 5)

    pillars = [
        'Hard Skills (kỹ năng tay làm)',
        'Soft Skills (giao tiếp, lắng nghe)',
        'Domain Knowledge (kiến thức ngành)',
        'Tech Literacy (web/mobile/API/DB)',
    ]
    for i, p in enumerate(pillars, 1):
        row = 24 + i
        ws.cell(row=row, column=1, value=p)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')

    dv2 = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
    dv2.add(f'B25:C28')
    ws.add_data_validation(dv2)

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22

    save(wb, 'TPL-001 — Skill Matrix Tự chấm điểm (Buổi 1).xlsx')


# ============================================================
def build_tpl_006():
    """Stakeholder Matrix — Power × Interest sortable."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Stakeholder Matrix'
    add_brand_header(ws, 'TPL-006', 'Buổi 3')

    add_title(ws, 'TPL-006 — Stakeholder Power × Interest Matrix',
              'Phân loại bên liên quan để biết ai cần chăm sóc kỹ.', 'G1')

    headers = ['#', 'Tên / Vai trò', 'Phòng ban', 'Power', 'Interest', 'Nhóm', 'Cách giao tiếp', 'Tần suất']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))

    # Sample rows
    samples = [
        (1, 'Nguyễn Văn A — TGĐ', 'Ban TGĐ', 'Cao', 'Cao', 'Quản lý sát', 'Họp 1-1', '2 tuần/lần'),
        (2, 'Trần Thị B — QL Kho', 'IT', 'Cao', 'Cao', 'Quản lý sát', 'Họp + chat', 'Hàng ngày'),
        (3, 'Lê Văn C — KTT', 'TC', 'Cao', 'Thấp', 'Giữ hài lòng', 'Email update', 'Tháng'),
        (4, 'Phạm D — NV kho', 'IT', 'Thấp', 'Cao', 'Giữ thông tin', 'Demo cuối sprint', '2 tuần'),
        (5, 'Tr.phòng MH', 'MH', 'Thấp', 'Thấp', 'Theo dõi', 'Newsletter chung', 'Quý'),
    ]
    for i, row in enumerate(samples, 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            style_data_cell(cell, italic=True)

    # Empty rows for fill
    for i in range(10, 30):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=c)
            cell.border = thin_border

    # Data validations
    dv_pi = DataValidation(type='list', formula1='"Cao,Thấp"', allow_blank=True)
    dv_pi.add('D5:E30')
    ws.add_data_validation(dv_pi)

    dv_group = DataValidation(type='list',
        formula1='"Quản lý sát,Giữ hài lòng,Giữ thông tin,Theo dõi"', allow_blank=True)
    dv_group.add('F5:F30')
    ws.add_data_validation(dv_group)

    # Conditional formatting cho Nhóm
    ws.conditional_formatting.add('F5:F30',
        CellIsRule(operator='equal', formula=['"Quản lý sát"'],
                   fill=PatternFill('solid', fgColor='D41B3F'),
                   font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add('F5:F30',
        CellIsRule(operator='equal', formula=['"Giữ hài lòng"'],
                   fill=PatternFill('solid', fgColor='F4A460'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add('F5:F30',
        CellIsRule(operator='equal', formula=['"Giữ thông tin"'],
                   fill=PatternFill('solid', fgColor='87CEEB'),
                   font=Font(bold=True)))

    # Column widths
    widths = [6, 30, 14, 10, 10, 18, 28, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = 'A5'

    # 4 quadrant guide
    ws.cell(row=33, column=1, value='4 QUADRANT — CÁCH TIẾP CẬN').font = subheader_font
    ws.cell(row=33, column=1).fill = subheader_fill
    ws.merge_cells('A33:H33')

    quad_headers = ['Quadrant', 'Power', 'Interest', 'Cách tiếp cận', '% thời gian giao tiếp']
    for c, h in enumerate(quad_headers, 1):
        ws.cell(row=34, column=c, value=h)
    style_header_row(ws, 34, len(quad_headers))

    quads = [
        ('Quản lý sát', 'Cao', 'Cao', 'Họp 1-1 · báo cáo tuần · xin ý kiến mọi quyết định lớn', '80%'),
        ('Giữ hài lòng', 'Cao', 'Thấp', 'Update khi có cột mốc lớn · không làm phiền chi tiết', '10%'),
        ('Giữ thông tin', 'Thấp', 'Cao', 'Newsletter · demo định kỳ · thu feedback', '8%'),
        ('Theo dõi', 'Thấp', 'Thấp', 'Cập nhật tổng quát · không tốn nhiều thời gian', '2%'),
    ]
    for i, q in enumerate(quads, 35):
        for c, val in enumerate(q, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = thin_border
            cell.alignment = cell_align

    save(wb, 'TPL-006 — Stakeholder Power-Interest Matrix (Buổi 3).xlsx')


# ============================================================
def build_tpl_007():
    """Lifecycle Tracking — backlog với 7 trạng thái."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Backlog Tracking'
    add_brand_header(ws, 'TPL-007', 'Buổi 3 + 9')

    add_title(ws, 'TPL-007 — Lifecycle Tracking Backlog',
              'Theo dõi từng yêu cầu xuyên 7 trạng thái — Phát hiện → Đóng.', 'K1')

    headers = ['ID', 'Tên yêu cầu', 'Người đề xuất', 'BA phụ trách', 'Trạng thái',
               'Sprint', 'Ngày tạo', 'Ngày update', 'Ưu tiên', 'Note']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))

    samples = [
        ('FR-001', 'Quét QR cấp phát', 'QL Kho', 'Nguyễn A', '04. Triển khai', 'Sprint 5',
         '01/03/2026', '14/03/2026', 'MUST', 'Dev đang code'),
        ('FR-002', 'Báo cáo tồn kho', 'CFO', 'Nguyễn A', '03. Duyệt', 'Sprint 6',
         '05/03/2026', '12/03/2026', 'SHOULD', 'Đợi sprint sau'),
        ('FR-003', 'Cảnh báo license', 'QL Kho', 'Nguyễn A', '02. Phân tích', 'Sprint 7',
         '10/03/2026', '13/03/2026', 'COULD', 'Cần spec ngày X'),
    ]
    for i, row in enumerate(samples, 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            style_data_cell(cell, italic=True)

    # Empty rows
    for i in range(8, 50):
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).border = thin_border

    # DV trạng thái
    dv_status = DataValidation(type='list',
        formula1='"01. New,02. Phân tích,03. Duyệt,04. Triển khai,05. Verified,06. Closed,07. Rejected"',
        allow_blank=True)
    dv_status.add('E5:E50')
    ws.add_data_validation(dv_status)

    # DV ưu tiên
    dv_pri = DataValidation(type='list', formula1='"MUST,SHOULD,COULD,WON\'T"', allow_blank=True)
    dv_pri.add('I5:I50')
    ws.add_data_validation(dv_pri)

    # Color code priority
    ws.conditional_formatting.add('I5:I50',
        CellIsRule(operator='equal', formula=['"MUST"'],
                   fill=PatternFill('solid', fgColor='D41B3F'),
                   font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add('I5:I50',
        CellIsRule(operator='equal', formula=['"SHOULD"'],
                   fill=PatternFill('solid', fgColor='F4A460')))
    ws.conditional_formatting.add('I5:I50',
        CellIsRule(operator='equal', formula=['"COULD"'],
                   fill=PatternFill('solid', fgColor='87CEEB')))

    widths = [12, 30, 16, 16, 22, 12, 14, 14, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A5'

    # 7 Trạng thái guide
    ws.cell(row=52, column=1, value='7 TRẠNG THÁI CHUẨN').font = subheader_font
    ws.cell(row=52, column=1).fill = subheader_fill
    ws.merge_cells('A52:J52')

    statuses = [
        ('01', 'New / Identified', 'Mới được nêu ra, chưa làm gì'),
        ('02', 'Elaborated', 'BA đã mô tả chi tiết, có acceptance'),
        ('03', 'Approved', 'Sponsor duyệt, sẵn sàng vào sprint'),
        ('04', 'In Progress', 'Dev đang code, QA đang test'),
        ('05', 'Verified', 'Test pass, demo OK'),
        ('06', 'Closed', 'Sign-off, deploy production'),
        ('07', 'Rejected / Deferred', 'Bị từ chối hoặc hoãn phase sau'),
    ]
    for c, h in enumerate(['Code', 'Tên', 'Ý nghĩa'], 1):
        ws.cell(row=53, column=c, value=h)
    style_header_row(ws, 53, 3)
    for i, s in enumerate(statuses, 54):
        for c, val in enumerate(s, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = thin_border
            cell.alignment = cell_align

    save(wb, 'TPL-007 — Backlog Tracking 7 trạng thái (Buổi 3+9).xlsx')


# ============================================================
def build_tpl_014():
    """User Story Map — 2D matrix."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Story Map'
    add_brand_header(ws, 'TPL-014', 'Buổi 5')

    add_title(ws, 'TPL-014 — User Story Map 2D',
              'Trục ngang = hành trình user. Trục dọc = priority. MVP = MUST.', 'G1')

    # 5 phase columns
    phases = ['Tìm hiểu', 'Đặt lịch', 'Theo dõi', 'Trải nghiệm dịch vụ', 'Sau dịch vụ']
    ws.cell(row=4, column=1, value='Priority \\ Phase')
    for c, p in enumerate(phases, 2):
        ws.cell(row=4, column=c, value=p)
    style_header_row(ws, 4, len(phases) + 1)

    # 3 priority rows + MVP line
    priorities = ['MUST (MVP)', 'SHOULD', 'COULD']
    colors = ['D41B3F', 'F4A460', '87CEEB']
    for i, (pri, color) in enumerate(zip(priorities, colors), 5):
        cell = ws.cell(row=i, column=1, value=pri)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(bold=True, color=WHITE if color == 'D41B3F' else BRAND_DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        # Empty cells to fill with US IDs
        for c in range(2, len(phases) + 2):
            sample = '' if i > 5 else f'US-{(c-2)*10 + 1:03d}, US-{(c-2)*10 + 2:03d}'
            cell = ws.cell(row=i, column=c, value=sample)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = thin_border
            cell.font = placeholder_font if sample else cell_font
        ws.row_dimensions[i].height = 80

    # MVP line note
    ws.cell(row=8, column=1, value='──────── MVP line ──────────').font = Font(italic=True, color='808080')
    ws.merge_cells('A8:F8')

    # Column widths
    ws.column_dimensions['A'].width = 18
    for c in range(2, len(phases) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 25

    # Hướng dẫn dùng
    ws.cell(row=11, column=1, value='HƯỚNG DẪN').font = subheader_font
    ws.cell(row=11, column=1).fill = subheader_fill
    ws.merge_cells('A11:F11')
    instructions = [
        '1. Cột (phase) = giai đoạn user journey theo thời gian — không phải module',
        '2. Hàng (priority) = mức độ ưu tiên (MUST/SHOULD/COULD)',
        '3. Cell chứa US ID — vd. US-001, US-002',
        '4. Hàng MUST = MVP (release đầu tiên)',
        '5. Move sticky thoải mái — không "lock" vào cell',
        '6. Update mỗi sprint review',
    ]
    for i, ins in enumerate(instructions, 12):
        ws.cell(row=i, column=1, value=ins)
        ws.merge_cells(f'A{i}:F{i}')

    # US Detail sheet
    ws2 = wb.create_sheet('US Detail')
    add_title(ws2, 'Chi tiết User Story', 'Fill chi tiết mỗi US ở đây — link với sheet Story Map.', 'F1')
    headers = ['US ID', 'Phase', 'Priority', 'Title', 'Acceptance', 'Sprint']
    for c, h in enumerate(headers, 1):
        ws2.cell(row=4, column=c, value=h)
    style_header_row(ws2, 4, len(headers))
    for i in range(5, 30):
        for c in range(1, len(headers) + 1):
            ws2.cell(row=i, column=c).border = thin_border
    widths = [12, 18, 12, 35, 40, 10]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    dv = DataValidation(type='list', formula1='"MUST,SHOULD,COULD"', allow_blank=True)
    dv.add('C5:C30')
    ws2.add_data_validation(dv)

    save(wb, 'TPL-014 — User Story Map 2D (Buổi 5).xlsx')


# ============================================================
def build_tpl_019():
    """Test Cases — bảng nhiều row."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Test Cases'
    add_brand_header(ws, 'TPL-019', 'Buổi 7')

    add_title(ws, 'TPL-019 — Test Cases Template',
              'Mỗi US có 5-7 TC (positive + negative + edge case).', 'L1')

    headers = ['TC ID', 'Linked US', 'Linked UC', 'Type', 'Priority', 'Pre-condition',
               'Steps', 'Expected Result', 'Actual', 'Status', 'Tester', 'Test date', 'Note']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))

    samples = [
        ('TC-042-01', 'US-042', 'UC-005', 'Positive', 'High',
         'NV kho đã login. Có MAC-001 status Available.',
         '1. Click Quét QR\n2. Quét MAC-001\n3. Chọn Nguyễn A\n4. Confirm',
         'Tạo bản ghi · Email gửi 10s · Toast success', '', '', '', '', ''),
        ('TC-042-02', 'US-042', 'UC-005', 'Negative', 'High',
         'NV kho đã login. QR FAKE-QR-999 không có trong hệ thống.',
         '1. Click Quét QR\n2. Quét FAKE-QR-999',
         'Toast lỗi "QR không hợp lệ" · Không tạo record', '', '', '', '', ''),
    ]
    for i, row in enumerate(samples, 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            style_data_cell(cell, italic=True)

    for i in range(7, 50):
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).border = thin_border

    # DV Type
    dv_type = DataValidation(type='list', formula1='"Positive,Negative,Edge case"', allow_blank=True)
    dv_type.add('D5:D50')
    ws.add_data_validation(dv_type)

    dv_pri = DataValidation(type='list', formula1='"High,Medium,Low"', allow_blank=True)
    dv_pri.add('E5:E50')
    ws.add_data_validation(dv_pri)

    dv_status = DataValidation(type='list', formula1='"Pass,Fail,Block,N/A"', allow_blank=True)
    dv_status.add('J5:J50')
    ws.add_data_validation(dv_status)

    # Color code Type
    ws.conditional_formatting.add('D5:D50',
        CellIsRule(operator='equal', formula=['"Positive"'],
                   fill=PatternFill('solid', fgColor='C6EFCE')))
    ws.conditional_formatting.add('D5:D50',
        CellIsRule(operator='equal', formula=['"Negative"'],
                   fill=PatternFill('solid', fgColor='FFC7CE')))
    ws.conditional_formatting.add('D5:D50',
        CellIsRule(operator='equal', formula=['"Edge case"'],
                   fill=PatternFill('solid', fgColor='FFEB9C')))

    ws.conditional_formatting.add('J5:J50',
        CellIsRule(operator='equal', formula=['"Pass"'],
                   fill=PatternFill('solid', fgColor='C6EFCE'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add('J5:J50',
        CellIsRule(operator='equal', formula=['"Fail"'],
                   fill=PatternFill('solid', fgColor='FFC7CE'),
                   font=Font(bold=True)))

    widths = [12, 12, 12, 12, 10, 30, 35, 35, 30, 10, 14, 12, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'C5'

    # Summary stats
    ws.cell(row=52, column=1, value='SUMMARY').font = subheader_font
    ws.cell(row=52, column=1).fill = subheader_fill
    ws.merge_cells('A52:E52')
    ws.cell(row=53, column=1, value='Total TC').font = Font(bold=True)
    ws.cell(row=53, column=2, value='=COUNTA(A5:A50)')
    ws.cell(row=54, column=1, value='Pass').font = Font(bold=True, color='008000')
    ws.cell(row=54, column=2, value='=COUNTIF(J5:J50,"Pass")')
    ws.cell(row=55, column=1, value='Fail').font = Font(bold=True, color='C00000')
    ws.cell(row=55, column=2, value='=COUNTIF(J5:J50,"Fail")')
    ws.cell(row=56, column=1, value='Pass rate').font = Font(bold=True)
    ws.cell(row=56, column=2, value='=IFERROR(B54/B53,0)')
    ws.cell(row=56, column=2).number_format = '0%'

    save(wb, 'TPL-019 — Test Cases (Buổi 7).xlsx')


# ============================================================
def build_tpl_023():
    """Validation Checklist 10 mục + auto pass-rate."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Validation Checklist'
    add_brand_header(ws, 'TPL-023', 'Buổi 8')

    add_title(ws, 'TPL-023 — Validation Checklist (10 tiêu chí)',
              'Tích đủ 10 = SRS sẵn sàng cho dev.', 'E1')

    headers = ['#', 'Tiêu chí', 'Status', 'Section/Evidence', 'Note']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))

    criteria = [
        'Có ID rõ cho mỗi yêu cầu (FR-001, NFR-002…)',
        'Mỗi câu = 1 yêu cầu, không nhồi 2-3 thứ',
        'Tránh từ mơ hồ ("nhanh", "thân thiện") — đo bằng số cụ thể',
        'Có Acceptance Criteria Given-When-Then cho mỗi US',
        'NFR đầy đủ 4 nhóm: Performance, Security, Usability, Reliability',
        'Có tham chiếu wireframe / mockup / ERD link Figma',
        'Glossary đã định nghĩa thuật ngữ chuyên ngành',
        'Versioning + change log đã update',
        'Đã review với ít nhất 1 BA khác (peer review)',
        'Đã có sign-off từ stakeholder cấp quyết định',
    ]
    for i, c in enumerate(criteria, 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=c)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
        ws.row_dimensions[row].height = 30

    # DV Status
    dv = DataValidation(type='list', formula1='"Pass,Partial,Fail,N/A"', allow_blank=True)
    dv.add('C5:C14')
    ws.add_data_validation(dv)

    ws.conditional_formatting.add('C5:C14',
        CellIsRule(operator='equal', formula=['"Pass"'],
                   fill=PatternFill('solid', fgColor='C6EFCE'),
                   font=Font(bold=True, color='006100')))
    ws.conditional_formatting.add('C5:C14',
        CellIsRule(operator='equal', formula=['"Partial"'],
                   fill=PatternFill('solid', fgColor='FFEB9C'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add('C5:C14',
        CellIsRule(operator='equal', formula=['"Fail"'],
                   fill=PatternFill('solid', fgColor='FFC7CE'),
                   font=Font(bold=True, color='9C0006')))

    # Summary
    ws.cell(row=16, column=1, value='SUMMARY').font = subheader_font
    ws.cell(row=16, column=1).fill = subheader_fill
    ws.merge_cells('A16:E16')

    ws.cell(row=17, column=2, value='Pass count')
    ws.cell(row=17, column=3, value='=COUNTIF(C5:C14,"Pass")').font = Font(bold=True, color='008000')
    ws.cell(row=18, column=2, value='Partial count')
    ws.cell(row=18, column=3, value='=COUNTIF(C5:C14,"Partial")').font = Font(bold=True, color='9C5700')
    ws.cell(row=19, column=2, value='Fail count')
    ws.cell(row=19, column=3, value='=COUNTIF(C5:C14,"Fail")').font = Font(bold=True, color='C00000')
    ws.cell(row=20, column=2, value='Pass rate').font = Font(bold=True)
    pr = ws.cell(row=20, column=3, value='=COUNTIF(C5:C14,"Pass")/10')
    pr.font = Font(bold=True, size=14, color=BRAND_RED)
    pr.number_format = '0%'

    ws.cell(row=22, column=2, value='Decision:').font = Font(bold=True)
    ws.cell(row=22, column=3, value='=IF(C20=1,"✓ READY for dev",IF(C20>=0.8,"⚠ Cần fix 1-2 mục","✗ Cần fix lớn"))')
    ws.cell(row=22, column=3).font = Font(bold=True, size=12)

    widths = [6, 50, 14, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    save(wb, 'TPL-023 — Validation Checklist 10 tiêu chí (Buổi 8).xlsx')


# ============================================================
def build_tpl_026():
    """Impact Analysis — 5 dimension."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Impact Analysis'
    add_brand_header(ws, 'TPL-026', 'Buổi 9')

    add_title(ws, 'TPL-026 — Impact Analysis (5 dimension)',
              'Phân tích impact của 1 Change Request — không miss dimension nào.', 'F1')

    # Header info
    ws.cell(row=4, column=1, value='CR ID:').font = Font(bold=True)
    ws.cell(row=4, column=2, value='[ CR-XXX ]').font = placeholder_font
    ws.cell(row=4, column=4, value='Date:').font = Font(bold=True)
    ws.cell(row=4, column=5, value='__/__/____').font = placeholder_font
    ws.cell(row=5, column=1, value='Phân tích bởi:').font = Font(bold=True)
    ws.cell(row=5, column=2, value='[ BA + Tech Lead ]').font = placeholder_font

    # 5 Dimension
    sections = [
        ('1. 📄 TÀI LIỆU bị ảnh hưởng',
         ['Tài liệu', 'Section', 'Action', 'Effort (giờ)'],
         [('SRS Phân hệ X', '3.2.1, 3.2.5', 'Update + thêm flow', 2),
          ('UC-001', 'Main flow + Alt', 'Add alt flow', 1),
          ('US-005', 'Acceptance', 'Add 2 acceptance', 0.5),
          ('Wireframe', 'Toàn bộ', 'Redesign 2 màn', 4),
          ('Test cases TC-005', '8 cases', 'Update 5 + thêm 3', 2),]),
        ('2. 💻 CODE bị ảnh hưởng',
         ['Module', 'Action', 'Files', 'Effort (giờ)', 'Refactor?'],
         [('auth-service', 'Add OTP integration', 5, 16, 'Nhỏ'),
          ('frontend-login', 'Add option UI', 3, 8, 'Không'),
          ('notification-service', 'Modify support', 2, 4, 'Không'),]),
        ('3. 📅 SCHEDULE',
         ['Sprint', 'Impact', 'Action'],
         [('Sprint hiện tại', '+3 ngày', 'Pull US-008 ra, add US-005a'),
          ('Sprint sau', '-1 US', 'Sẽ overload nếu add lại'),
          ('Go-live deadline', 'Không ảnh hưởng', 'Buffer còn 2 tuần'),]),
        ('4. 💰 BUDGET',
         ['Mục', 'Cost', 'Note'],
         [('Dev hours', '14tr', 'Fit contingency'),
          ('QA hours', '2.4tr', 'Fit contingency'),
          ('License Zalo OA', '5tr/year', 'Cost mới'),
          ('Tổng', '21.4tr', 'Còn budget contingency: 50tr → OK'),]),
        ('5. ⚠ RISK',
         ['Rủi ro', 'Likelihood', 'Impact', 'Mitigation'],
         [('Zalo OA approval mất 2 tuần', 'High', 'Medium', 'Apply ngay (đã apply 13/03)'),
          ('Zalo API rate limit', 'Medium', 'Low', 'Implement retry + fallback'),
          ('User confused 2 option', 'Medium', 'Low', 'UX research + A/B test'),]),
    ]

    current_row = 7
    for section_title, sec_headers, rows in sections:
        ws.cell(row=current_row, column=1, value=section_title).font = Font(bold=True, size=13, color=BRAND_RED)
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1

        for c, h in enumerate(sec_headers, 1):
            ws.cell(row=current_row, column=c, value=h)
        style_header_row(ws, current_row, len(sec_headers))
        current_row += 1

        for row in rows:
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=c, value=val)
                style_data_cell(cell, italic=True)
            current_row += 1
        # Empty rows for additional fill
        for _ in range(2):
            for c in range(1, len(sec_headers) + 1):
                ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1
        current_row += 1  # spacer

    # Recommendation
    ws.cell(row=current_row, column=1, value='RECOMMENDATION').font = Font(bold=True, size=14, color=BRAND_RED)
    ws.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 1
    rec = [
        ('Value', '⭐⭐⭐⭐ Cao'),
        ('Cost', '21.4tr — Reasonable'),
        ('Effort', '28h dev + 6h QA'),
        ('Risk', 'Medium — Có mitigation'),
        ('Decision recommend', 'APPROVE'),
    ]
    for k, v in rec:
        ws.cell(row=current_row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=v)
        current_row += 1

    widths = [25, 25, 18, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    save(wb, 'TPL-026 — Impact Analysis 5 dimension (Buổi 9).xlsx')


# ============================================================
def build_tpl_027():
    """Traceability Matrix — 5 cột BR → TC."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Traceability'
    add_brand_header(ws, 'TPL-027', 'Buổi 9 + 11')

    add_title(ws, 'TPL-027 — Traceability Matrix',
              'Truy vết từ Business → Test xuyên 5 cấp.', 'G1')

    headers = ['#', 'BR ID', 'BR Description', 'US ID', 'FR ID', 'UC ID', 'TC IDs']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))

    samples = [
        (1, 'BR-001', 'Tăng 30% lượng đặt online', 'US-042', 'FR-018', 'UC-005', 'TC-042-01..08'),
        (2, 'BR-001', '(cùng BR)', 'US-043', 'FR-019', 'UC-006', 'TC-043-01..05'),
        (3, 'BR-002', 'Giảm 50% thời gian admin', 'US-044', 'FR-020', 'UC-007', 'TC-044-01..06'),
        (4, 'BR-002', '(cùng BR)', 'US-045', 'FR-020, FR-021', 'UC-008', 'TC-045-01..04'),
        (5, 'BR-003', 'NPS ≥ 8/10', 'US-046', 'FR-022', 'UC-009', 'TC-046-01..07'),
    ]
    for i, row in enumerate(samples, 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            style_data_cell(cell, italic=True)

    for i in range(10, 50):
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).border = thin_border

    widths = [6, 14, 35, 14, 14, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A5'

    # Statistics
    ws.cell(row=52, column=1, value='STATISTICS').font = subheader_font
    ws.cell(row=52, column=1).fill = subheader_fill
    ws.merge_cells('A52:G52')
    stats = [
        ('Số lượng BR', '=SUMPRODUCT((B5:B50<>"")/COUNTIF(B5:B50,B5:B50&""))'),
        ('Số lượng US', '=COUNTA(D5:D50)'),
        ('Số lượng FR', '=COUNTA(E5:E50)'),
        ('Số lượng UC', '=COUNTA(F5:F50)'),
    ]
    for i, (label, formula) in enumerate(stats, 53):
        ws.cell(row=i, column=2, value=label).font = Font(bold=True)
        ws.cell(row=i, column=3, value=formula).font = Font(bold=True, color=BRAND_RED)

    # Coverage check
    ws.cell(row=58, column=1, value='COVERAGE CHECK').font = subheader_font
    ws.cell(row=58, column=1).fill = subheader_fill
    ws.merge_cells('A58:G58')
    checks = [
        'Mỗi BR có ≥ 1 US?',
        'Mỗi US có ≥ 1 FR?',
        'Mỗi FR có ≥ 1 UC (nếu phức tạp)?',
        'Mỗi US/UC có ≥ 3 TC (positive + negative + edge)?',
    ]
    for i, c in enumerate(checks, 59):
        ws.cell(row=i, column=1, value='☐')
        ws.cell(row=i, column=2, value=c)
        ws.merge_cells(f'B{i}:G{i}')

    save(wb, 'TPL-027 — Traceability Matrix 5 cấp (Buổi 9).xlsx')


# ============================================================
def build_tpl_028():
    """Prioritization — RICE formula."""
    wb = Workbook()

    # Sheet 1: MoSCoW
    ws1 = wb.active
    ws1.title = 'MoSCoW'
    add_brand_header(ws1, 'TPL-028', 'Buổi 9')

    add_title(ws1, 'TPL-028 — MoSCoW Prioritization',
              'Phân loại 4 nhóm: MUST / SHOULD / COULD / WON\'T.', 'D1')

    headers = ['US ID', 'Tên User Story', 'MoSCoW', 'Lý do']
    for c, h in enumerate(headers, 1):
        ws1.cell(row=4, column=c, value=h)
    style_header_row(ws1, 4, len(headers))

    samples = [
        ('US-042', 'Quét QR cấp phát', 'MUST', 'Core feature — không có = không launch được'),
        ('US-043', 'Xem lịch sử cấp phát', 'MUST', 'Audit trail bắt buộc'),
        ('US-044', 'Notification email', 'SHOULD', 'Quan trọng nhưng có workaround'),
        ('US-045', 'Báo cáo tồn kho', 'SHOULD', 'Manager cần — tháng đầu có thể manual'),
        ('US-046', 'Cảnh báo license hết hạn', 'COULD', 'Nice to have — phase 2 OK'),
        ('US-047', 'Voice search thiết bị', "WON'T", 'Phase sau, không cần phase này'),
    ]
    for i, row in enumerate(samples, 5):
        for c, val in enumerate(row, 1):
            cell = ws1.cell(row=i, column=c, value=val)
            style_data_cell(cell, italic=True)

    for i in range(11, 30):
        for c in range(1, len(headers) + 1):
            ws1.cell(row=i, column=c).border = thin_border

    dv = DataValidation(type='list', formula1='"MUST,SHOULD,COULD,WON\'T"', allow_blank=True)
    dv.add('C5:C30')
    ws1.add_data_validation(dv)

    # Color code MoSCoW
    ws1.conditional_formatting.add('C5:C30',
        CellIsRule(operator='equal', formula=['"MUST"'],
                   fill=PatternFill('solid', fgColor='D41B3F'),
                   font=Font(color=WHITE, bold=True)))
    ws1.conditional_formatting.add('C5:C30',
        CellIsRule(operator='equal', formula=['"SHOULD"'],
                   fill=PatternFill('solid', fgColor='F4A460'),
                   font=Font(bold=True)))
    ws1.conditional_formatting.add('C5:C30',
        CellIsRule(operator='equal', formula=['"COULD"'],
                   fill=PatternFill('solid', fgColor='87CEEB')))

    # Stats
    ws1.cell(row=32, column=1, value='DISTRIBUTION').font = subheader_font
    ws1.cell(row=32, column=1).fill = subheader_fill
    ws1.merge_cells('A32:D32')
    for i, m in enumerate(['MUST', 'SHOULD', 'COULD', "WON'T"], 33):
        ws1.cell(row=i, column=1, value=m).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=f'=COUNTIF(C5:C30,"{m}")').alignment = Alignment(horizontal='center')

    widths = [12, 35, 12, 35]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2: RICE
    ws2 = wb.create_sheet('RICE')
    add_title(ws2, 'RICE Score = (Reach × Impact × Confidence) / Effort',
              'Đo backlog bằng số objective. Sort descending → ưu tiên cao trước.', 'H1')

    headers2 = ['US ID', 'Tên', 'Reach (users/qtr)', 'Impact (0.25-3)',
                'Confidence (%)', 'Effort (weeks)', 'RICE Score', 'Note']
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=c, value=h)
    style_header_row(ws2, 4, len(headers2))

    samples2 = [
        ('US-042', 'Quét QR cấp phát', 500, 3, 90, 1, '=C5*D5*(E5/100)/F5', 'High value'),
        ('US-046', 'Cảnh báo license', 50, 2, 70, 0.5, '=C6*D6*(E6/100)/F6', 'Phase 2'),
        ('US-047', 'Voice search', 5000, 0.25, 30, 3, '=C7*D7*(E7/100)/F7', 'Low confidence'),
        ('US-048', 'Báo cáo Power BI', 100, 1, 80, 2, '=C8*D8*(E8/100)/F8', 'Manager request'),
    ]
    for i, row in enumerate(samples2, 5):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=c, value=val)
            style_data_cell(cell, italic=True if isinstance(val, str) and not val.startswith('=') else False)
            if c == 7:  # RICE score
                cell.font = Font(bold=True, color=BRAND_RED)
                cell.number_format = '0.0'

    for i in range(9, 30):
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=i, column=c)
            cell.border = thin_border
            if c == 7:
                cell.value = f'=IFERROR(C{i}*D{i}*(E{i}/100)/F{i},"")'
                cell.font = Font(bold=True, color=BRAND_RED)
                cell.number_format = '0.0'

    # Data bar conditional
    ws2.conditional_formatting.add('G5:G30',
        DataBarRule(start_type='min', end_type='max', color=BRAND_RED))

    widths = [12, 30, 16, 14, 14, 12, 14, 25]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # RICE guide
    ws2.cell(row=32, column=1, value='HƯỚNG DẪN CHẤM').font = subheader_font
    ws2.cell(row=32, column=1).fill = subheader_fill
    ws2.merge_cells('A32:H32')
    guides = [
        ('Reach', 'Bao nhiêu user dùng tính năng/quarter (số tuyệt đối)'),
        ('Impact', '3=Massive · 2=High · 1=Medium · 0.25=Low'),
        ('Confidence', '100% = có data · 80% = research · 50% = đoán có cơ sở · 20% = đoán hoàn toàn'),
        ('Effort', 'Person-weeks dev (1 = 1 dev × 1 tuần)'),
    ]
    for i, (k, v) in enumerate(guides, 33):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=v)
        ws2.merge_cells(f'B{i}:H{i}')

    save(wb, 'TPL-028 — MoSCoW + RICE Prioritization (Buổi 9).xlsx')


# ============================================================
def build_tpl_033():
    """Rubric chấm bài tập cuối khoá — 100 điểm."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rubric'
    add_brand_header(ws, 'TPL-033', 'Buổi 11')

    add_title(ws, 'TPL-033 — Rubric chấm Bài tập Cuối khoá (100 điểm)',
              'Pass ≥ 70 = đạt chứng nhận + 36 PDU.', 'F1')

    # Header info
    ws.cell(row=4, column=1, value='Tên học viên:').font = Font(bold=True)
    ws.cell(row=4, column=2, value='[ ... ]').font = placeholder_font
    ws.cell(row=4, column=4, value='Ngày chấm:').font = Font(bold=True)
    ws.cell(row=4, column=5, value='__/__/____').font = placeholder_font
    ws.cell(row=5, column=1, value='Instructor:').font = Font(bold=True)
    ws.cell(row=5, column=2, value='Trần Đức').font = placeholder_font

    # Rubric headers
    headers = ['#', 'Khu vực', 'Tiêu chí', 'Max', 'Đạt', 'Note']
    for c, h in enumerate(headers, 1):
        ws.cell(row=7, column=c, value=h)
    style_header_row(ws, 7, len(headers))

    rubric_items = [
        ('1', 'Planning', 'Mục tiêu kinh doanh rõ + đo bằng số', 3),
        ('1', 'Planning', 'Stakeholder matrix Power/Interest đầy đủ', 3),
        ('1', 'Planning', 'Timeline 1 tuần chi tiết theo ngày', 2),
        ('1', 'Planning', 'Deliverable list rõ + scope', 2),
        ('2', 'Elicitation', '3 buổi phỏng vấn với 3 vai trò khác nhau', 3),
        ('2', 'Elicitation', 'Pain points cụ thể (≥ 3/buổi)', 3),
        ('2', 'Elicitation', 'Yêu cầu nêu ra (≥ 5/buổi)', 3),
        ('2', 'Elicitation', 'AS-IS quy trình mô tả', 2),
        ('2', 'Elicitation', 'Áp dụng 5 Whys ≥ 1 lần', 2),
        ('2', 'Elicitation', 'Quote nguyên văn (≥ 1/buổi)', 2),
        ('3', 'Analysis', '≥ 3 BPMN đúng shape', 5),
        ('3', 'Analysis', 'ERD đúng cardinality + ≥ 4 entity', 5),
        ('3', 'Analysis', '5 Use Case Spec đầy đủ template', 5),
        ('3', 'Analysis', 'Logic + ngữ pháp đúng', 3),
        ('3', 'Analysis', 'Tự vẽ (không copy từ Demo)', 2),
        ('4', 'Design', '3-5 wireframe cho màn chính', 4),
        ('4', 'Design', 'Áp dụng 6 nguyên tắc UX', 4),
        ('4', 'Design', 'Có 3 mức fidelity (low/mid/high)', 3),
        ('4', 'Design', 'Sử dụng v0.dev hoặc Figma', 2),
        ('4', 'Design', 'Self-check theo TPL-015', 2),
        ('5', 'Documentation', 'SRS đầy đủ 8 mục', 5),
        ('5', 'Documentation', '20 User Story chuẩn INVEST', 5),
        ('5', 'Documentation', 'Acceptance Criteria Given-When-Then', 4),
        ('5', 'Documentation', '50 Test Cases (positive/negative/edge)', 4),
        ('5', 'Documentation', '7 quy tắc viết tài liệu rõ', 4),
        ('5', 'Documentation', 'Glossary + Change log', 3),
        ('6', 'Traceability', 'Bảng 5 cột BR → TC đầy đủ', 5),
        ('6', 'Traceability', 'Coverage check pass', 3),
        ('6', 'Traceability', 'Format chuẩn (Excel/MD table)', 2),
        ('7', 'AI Applied', 'Có ≥ 4-5 prompt áp dụng', 2),
        ('7', 'AI Applied', 'Có verify output AI', 2),
        ('7', 'AI Applied', 'Note "AI nào, prompt gì"', 1),
    ]
    for i, (sec, area, criterion, mx) in enumerate(rubric_items, 8):
        ws.cell(row=i, column=1, value=sec).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=2, value=area)
        ws.cell(row=i, column=3, value=criterion)
        ws.cell(row=i, column=4, value=mx).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=5).alignment = Alignment(horizontal='center')  # empty for fill
        for c in range(1, 7):
            ws.cell(row=i, column=c).border = thin_border

    last_row = 7 + len(rubric_items)
    # Total row
    total_row = last_row + 1
    ws.cell(row=total_row, column=2, value='TỔNG').font = Font(bold=True, size=12)
    ws.cell(row=total_row, column=4, value=f'=SUM(D8:D{last_row})').font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=f'=SUM(E8:E{last_row})').font = Font(bold=True, size=12, color=BRAND_RED)
    for c in range(1, 7):
        ws.cell(row=total_row, column=c).fill = PatternFill('solid', fgColor=LIGHT_GREY)
        ws.cell(row=total_row, column=c).border = thin_border

    # Status
    status_row = total_row + 2
    ws.cell(row=status_row, column=2, value='STATUS').font = Font(bold=True, size=12)
    ws.cell(row=status_row, column=3, value=f'=IF(E{total_row}>=70,"✓ PASS — Đạt chứng nhận",IF(E{total_row}>=50,"⚠ RESUBMIT — Sửa + nộp lại","✗ FAIL — Học lại"))')
    ws.cell(row=status_row, column=3).font = Font(bold=True, size=14)
    ws.merge_cells(f'C{status_row}:F{status_row}')

    # Conditional formatting for E column (score)
    ws.conditional_formatting.add(f'E{total_row}',
        ColorScaleRule(start_type='num', start_value=0, start_color='FFC7CE',
                       mid_type='num', mid_value=70, mid_color='FFEB9C',
                       end_type='num', end_value=100, end_color='C6EFCE'))

    # Feedback section
    fb_row = status_row + 2
    ws.cell(row=fb_row, column=1, value='FEEDBACK').font = subheader_font
    ws.cell(row=fb_row, column=1).fill = subheader_fill
    ws.merge_cells(f'A{fb_row}:F{fb_row}')

    fb_items = [
        ('Điểm mạnh (3 điểm)', '[ ... ]'),
        ('Cần cải thiện (3 điểm)', '[ ... ]'),
        ('Action items', '[ ... ]'),
        ('Lời chúc', '[ ... ]'),
    ]
    for i, (k, v) in enumerate(fb_items, fb_row + 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v).font = placeholder_font
        ws.merge_cells(f'B{i}:F{i}')
        ws.row_dimensions[i].height = 40

    widths = [6, 20, 50, 8, 10, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    save(wb, 'TPL-033 — Rubric Bài tập Cuối khoá 100đ (Buổi 11).xlsx')


# ============================================================
if __name__ == '__main__':
    print('=== Building 10 Excel templates ===\n')
    build_tpl_001()
    build_tpl_006()
    build_tpl_007()
    build_tpl_014()
    build_tpl_019()
    build_tpl_023()
    build_tpl_026()
    build_tpl_027()
    build_tpl_028()
    build_tpl_033()
    print('\n=== HOÀN THÀNH 10 file Excel ===')
